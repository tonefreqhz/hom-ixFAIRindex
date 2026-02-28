import os
import re
import numpy as np
import matplotlib.pyplot as plt

import matplotlib.cm as cm
import matplotlib.colors as mcolors
import matplotlib.patches as mpatches

from matplotlib.animation import FuncAnimation, FFMpegWriter
from openpyxl import load_workbook


# -----------------------------
# SETTINGS
# -----------------------------
IN_XLSX = r"C:\Users\peewe\OneDrive\Desktop\homeix\housing market banded-country-England and Wales-by-district-old and newbuild-1995 to may 2022 (2).xlsx"
OUT_MP4 = "ew_sales_by_price_band_grand_total_1995_2022.mp4"

FPS = 1
SECONDS_PER_YEAR = 2
FRAME_REPEAT = FPS * SECONDS_PER_YEAR

DPI = 160
BITRATE = 3500

# Year tabs are usually named "1995", "1996", ...
SHEET_YEAR_REGEX = r"^\s*(19\d{2}|20\d{2})\s*$"


# -----------------------------
# HELPERS
# -----------------------------
def _norm(x) -> str:
    if x is None:
        return ""
    s = str(x).strip()
    s = re.sub(r"\s+", " ", s)
    return s


def parse_year_from_sheet(sheet_name: str):
    m = re.match(SHEET_YEAR_REGEX, str(sheet_name))
    return int(m.group(1)) if m else None


def to_number(x):
    """Convert Excel cell to float; handles commas and blanks."""
    if x is None:
        return np.nan
    if isinstance(x, (int, float)) and not isinstance(x, bool):
        return float(x)
    s = str(x).strip().replace(",", "")
    if s == "":
        return np.nan
    try:
        return float(s)
    except Exception:
        return np.nan


def detect_header_row_ws(ws, max_scan_rows: int = 80):
    """
    Find the header row by scoring rows containing:
    'Area', 'Type', 'Under 10,000', 'Total'
    and penalising the helper row like: 'top of sheet , under , Under 10,000 ...'
    """
    best_r, best_score = 1, -10**9

    max_r = min(max_scan_rows, ws.max_row)

    for r in range(1, max_r + 1):
        row_vals = [_norm(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)]
        row_l = [v.lower() for v in row_vals]
        nonempty = sum(1 for v in row_l if v and v != "nan")

        score = nonempty * 50
        if "area" in row_l:
            score += 5000
        if "type" in row_l:
            score += 5000
        if any(v == "under 10,000" for v in row_l):
            score += 9000
        if any(v == "total" for v in row_l):
            score += 4500

        # Penalize the 'under' helper row if it doesn't also contain Area & Type
        if "under" in row_l and ("area" not in row_l or "type" not in row_l):
            score -= 7000

        if score > best_score:
            best_r, best_score = r, score

    return best_r, best_score


def get_column_indices(ws, header_row: int):
    """
    From the header row, find:
      - Area column
      - Type column
      - Total column (RIGHTMOST 'Total' if multiple)
      - Band columns = columns between Type and Total
    """
    headers = []
    for c in range(1, ws.max_column + 1):
        headers.append(_norm(ws.cell(row=header_row, column=c).value))

    headers_l = [h.lower() for h in headers]

    def find_exact(name):
        for i, h in enumerate(headers_l, start=1):
            if h == name:
                return i
        return None

    area_col = find_exact("area")
    type_col = find_exact("type")

    total_candidates = [i for i, h in enumerate(headers_l, start=1) if h == "total"]
    total_col = max(total_candidates) if total_candidates else None

    if area_col is None or type_col is None or total_col is None:
        raise ValueError(
            f"Could not find Area/Type/Total in header row {header_row}. "
            f"Found Area={area_col}, Type={type_col}, Total={total_col}."
        )

    band_cols = list(range(type_col + 1, total_col))
    band_labels = [headers[c - 1] for c in band_cols]

    if not band_cols:
        raise ValueError("No band columns detected between Type and Total.")

    return area_col, type_col, total_col, band_cols, band_labels


def find_grand_total_row_by_max_total(ws, header_row: int, total_col: int):
    """
    Grand total row = row below header with the maximum numeric value in Total column.
    Works even if the 'Total' row moves (AD1698, AD1988, etc).
    """
    best_r = None
    best_total = -1.0

    for r in range(header_row + 1, ws.max_row + 1):
        v = to_number(ws.cell(row=r, column=total_col).value)
        if np.isfinite(v) and v > best_total:
            best_total = v
            best_r = r

    if best_r is None:
        raise ValueError("Couldn't find any numeric values in the Total column.")
    return best_r, best_total


def make_band_palette(bands):
    cmap = cm.get_cmap("turbo", len(bands))
    return {b: mcolors.to_hex(cmap(i)) for i, b in enumerate(bands)}


# -----------------------------
# MAIN
# -----------------------------
def main():
    if not os.path.exists(IN_XLSX):
        raise FileNotFoundError(f"Workbook not found:\n{IN_XLSX}")

    wb = load_workbook(IN_XLSX, data_only=True)

    year_sheets = []
    for s in wb.sheetnames:
        y = parse_year_from_sheet(s)
        if y is not None:
            year_sheets.append((y, s))
    year_sheets.sort(key=lambda t: t[0])

    if not year_sheets:
        raise ValueError("No year-named sheets found (e.g. '1995'). Edit SHEET_YEAR_REGEX if needed.")

    records = []
    bands_ref = None

    for year, sheet in year_sheets:
        ws = wb[sheet]

        header_row, score = detect_header_row_ws(ws, max_scan_rows=80)
        area_col, type_col, total_col, band_cols, band_labels = get_column_indices(ws, header_row)

        total_row, sheet_total = find_grand_total_row_by_max_total(ws, header_row, total_col)

        # Read band values from the chosen row
        vals = np.array([to_number(ws.cell(row=total_row, column=c).value) for c in band_cols], dtype=float)
        vals = np.nan_to_num(vals, nan=0.0)

        # Align band order across years (use first year as reference labels)
        if bands_ref is None:
            bands_ref = band_labels
        else:
            label_to_idx = {lab: i for i, lab in enumerate(band_labels)}
            aligned_vals = []
            for lab in bands_ref:
                aligned_vals.append(vals[label_to_idx[lab]] if lab in label_to_idx else 0.0)
            vals = np.array(aligned_vals, dtype=float)
            band_labels = bands_ref

        # Debug line: shows exactly what it's reading
        print(
            f"[sheet {sheet}] header_row={header_row} score={score} "
            f"TotalCol={total_col} TotalRow={total_row} grand_total={sheet_total:,.0f}"
        )

        records.append((year, sheet, band_labels, vals, sheet_total))

    bands = records[0][2]
    band_to_color = make_band_palette(bands)
    bar_colors = [band_to_color[b] for b in bands]
    legend_handles = [mpatches.Patch(color=band_to_color[b], label=str(b)) for b in bands]

    all_vals = np.vstack([r[3] for r in records])
    ymax = float(np.nanmax(all_vals)) if np.isfinite(all_vals).any() else 1.0
    ymax = max(1.0, ymax) * 1.12

    x = np.arange(len(bands))

    fig, ax = plt.subplots(figsize=(15, 7.5), dpi=DPI)
    fig.patch.set_facecolor("white")

    def draw(sheet, vals, sheet_total):
        ax.clear()

        ax.bar(x, vals, color=bar_colors, edgecolor="white", linewidth=0.6)

        ax.set_ylim(0, ymax)
        ax.set_title(
            "Volumes of Sales by Price Band — England & Wales (Grand Total per year)",
            loc="left",
            fontsize=14,
        )

        ax.text(
            0.99, 0.96,
            f"{sheet}\nTotal sales: {int(sheet_total):,}",
            transform=ax.transAxes,
            ha="right", va="top",
            fontsize=12,
            bbox=dict(facecolor="white", alpha=0.88, edgecolor="none", pad=4),
        )

        ax.set_ylabel("Number of sales")
        ax.set_xlabel("Price band (£)")
        ax.set_xticks(x)
        ax.set_xticklabels([str(b) for b in bands], rotation=45, ha="right")
        ax.grid(True, axis="y", alpha=0.25)

        ax.legend(
            handles=legend_handles,
            title="Price band key (£)",
            loc="upper left",
            bbox_to_anchor=(1.01, 1.0),
            frameon=True,
            fontsize=8.0,
            title_fontsize=9.0,
            ncol=1,
        )

        plt.tight_layout(rect=[0, 0, 0.78, 1])

    # Repeat frames so each year holds on screen for N seconds
    frame_map = []
    for i in range(len(records)):
        frame_map.extend([i] * FRAME_REPEAT)

    def animate(k):
        i = frame_map[k]
        _, sheet, _, vals, sheet_total = records[i]
        draw(sheet, vals, sheet_total)
        return []

    anim = FuncAnimation(fig, animate, frames=len(frame_map), interval=1000 // FPS, blit=False)

    writer = FFMpegWriter(fps=FPS, metadata={"artist": "homeix"}, bitrate=BITRATE)
    anim.save(OUT_MP4, writer=writer)

    plt.close(fig)
    print(f"\nSaved MP4:\n{os.path.abspath(OUT_MP4)}")


if __name__ == "__main__":
    main()
